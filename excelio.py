"""
XHS Shots — Excel I/O
======================
Read a column of Xiaohongshu links from an .xlsx file, run captures, and write
the resulting screenshots back into a chosen (or auto-created) column.

Handles two real-world cases that caused "No links found":
  * links stored as Excel HYPERLINKS (clickable, value may be a title, not a URL)
  * links living on a non-active sheet
"""

from __future__ import annotations

import re
import time
import urllib.parse
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from capture import CaptureEngine, LoginWallError, CaptureError, _col_to_idx, sanitize_filename


def _resolve_anchor(ws, row: int, col: int):
    """Return (row, col) of the top-left cell of the merge containing (row, col),
    or (row, col) itself if it is not merged. Needed because writing a value or
    anchoring an image to a MergedCell raises (read-only)."""
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row, rng.min_col
    return row, col


def _out_set(ws, row: int, col: int, value):
    r, c = _resolve_anchor(ws, row, col)
    ws.cell(row=r, column=c).value = value


def load_workbook(path: str | Path, read_only: bool = False, data_only: bool = True):
    return openpyxl.load_workbook(path, read_only=read_only, data_only=data_only)


def list_sheets(path: str | Path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


def _get_sheet(wb, sheet: Optional[str]):
    if sheet:
        return wb[sheet]
    return wb.active


_URL_RE = re.compile(r'https?://[^\s"\'()<>]+', re.IGNORECASE)


def _cell_url(ws_val, row: int, col: int, ws_fmt=None) -> Optional[str]:
    """Best-effort extract a URL from a cell.

    Tries, in order:
      1. an http(s) hyperlink target
      2. an http(s) cached *value*  (ws_val loaded with data_only=True)
      3. an http(s) URL mined out of a *formula* string
         (ws_fmt loaded with data_only=False — e.g. =HYPERLINK("http://..."))
      4. any non-empty value (fallback, so title-text links still surface)
    """
    cell = ws_val.cell(row=row, column=col)
    hl = None
    try:
        hl = cell.hyperlink  # not available in read_only mode
    except AttributeError:
        hl = None
    target = getattr(hl, "target", None) if hl else None
    if target and str(target).strip().lower().startswith("http"):
        return str(target).strip()

    v = cell.value
    if v is not None and str(v).strip().lower().startswith("http"):
        return str(v).strip()

    # formula fallback: the visible cached value may be empty (formula not
    # recalculated / not cached), but the formula string still holds the URL
    if ws_fmt is not None:
        f = ws_fmt.cell(row=row, column=col).value
        if isinstance(f, str):
            m = _URL_RE.search(f)
            if m:
                return m.group(0).rstrip(".,;")

    if v is not None and str(v).strip():
        return str(v).strip()
    return None


def preview_columns(path: str | Path, sheet: Optional[str] = None) -> list[dict]:
    """Return column metadata for the (chosen) sheet, showing URL/hyperlink sample.
    Loaded normally (not read_only) so Excel HYPERLINK targets are accessible."""
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    ws = _get_sheet(wb, sheet)
    cols = []
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        sample = ""
        first_any = ""
        for r in range(1, min(max_row, 8) + 1):
            u = _cell_url(ws, r, c)
            if not u:
                continue
            if not first_any:
                first_any = u
            if u.lower().startswith("http"):  # prefer a real link as the sample
                sample = u[:46]
                break
        if not sample:
            sample = first_any[:46]
        cols.append({"index": c, "letter": letter, "sample": sample})
    wb.close()
    return cols


def read_links(path: str | Path, link_col, header_rows: int = 1,
               sheet: Optional[str] = None) -> list[tuple[int, str]]:
    """Return [(row_index, url), ...] for data rows (skipping headers)."""
    wb_val = openpyxl.load_workbook(path, read_only=False, data_only=True)   # cached values + hyperlinks
    wb_fmt = openpyxl.load_workbook(path, read_only=False, data_only=False)  # formula strings
    ws = _get_sheet(wb_val, sheet)
    ws_fmt = _get_sheet(wb_fmt, sheet)
    col = _col_to_idx(link_col)
    out = []
    for r in range(header_rows + 1, (ws.max_row or 0) + 1):
        u = _cell_url(ws, r, col, ws_fmt)
        if u:
            out.append((r, u))
    wb_val.close()
    wb_fmt.close()
    return out


def _looks_like_link(s: str) -> bool:
    """Looser than strict http:// — catches xhs links without a scheme,
    short-domain links, etc."""
    s = str(s).strip()
    low = s.lower()
    if low.startswith("http") or "://" in low:
        return True
    if "xiaohongshu" in low or "xhslink" in low or "xhs.cn" in low:
        return True
    # a generic domain-ish token: word.word/...
    if re.search(r"[a-z0-9-]+\.[a-z]{2,}/", low):
        return True
    return False


def _normalize_and_validate_url(raw: str) -> Optional[str]:
    """Turn a cell value into a navigable URL, or return None if it can't be
    made into a valid http(s) URL.

    Handles the common Excel garbage that Playwright's page.goto rejects as
    'invalid URL':
      * missing scheme  -> prepend https://
      * stray whitespace / newlines inside the URL -> strip them
      * surrounding quotes / angle brackets
      * trailing punctuation (.,;)
      * bare 'https://' with nothing after it
    """
    s = str(raw).strip().strip('"').strip("'").strip("<>").strip()
    if not s:
        return None
    if "://" not in s:
        s = "https://" + s
    s = s.rstrip(".,;")

    def _ok(candidate: str) -> Optional[str]:
        try:
            p = urllib.parse.urlparse(candidate)
        except (ValueError, UnicodeError):
            return None
        if p.scheme not in ("http", "https"):
            return None
        if not p.netloc:
            return None
        # netloc must be a real host: no whitespace and contain a dot
        # (rejects 'https:// ', 'https://  xiaohongshu.com', 'https://12345')
        if " " in p.netloc or "\n" in p.netloc or "\t" in p.netloc:
            return None
        if "." not in p.netloc:
            return None
        # Return the canonicalized form — urlparse silently drops invalid chars
        # (e.g. a newline) from netloc, so the raw input may still contain junk
        # that page.goto would reject. Rebuilding from the parsed parts is safe.
        return urllib.parse.urlunparse(p)

    res = _ok(s)
    if res:
        return res
    # Recovery: remove ALL whitespace (catches 'https:// www.xiaohongshu.com')
    s2 = re.sub(r"\s+", "", s)
    if s2 != s:
        return _ok(s2)
    return None


def find_link_columns(path: str | Path, sheet: Optional[str] = None,
                      header_rows: int = 1, n_rows: int = 300,
                      all_sheets: bool = False) -> list[dict]:
    """Scan the (chosen or all) sheet(s) and report which columns actually
    contain link-like content (value / hyperlink / formula). Helps the user
    pick the right column when they selected the wrong one, or when the links
    live on another sheet."""
    wb_val = openpyxl.load_workbook(path, read_only=False, data_only=True)
    wb_fmt = openpyxl.load_workbook(path, read_only=False, data_only=False)
    sheets = wb_fmt.sheetnames if all_sheets else ([_get_sheet(wb_val, sheet).title] if sheet else [wb_val.active.title])
    hits = []
    for sname in sheets:
        ws = wb_val[sname]
        ws_fmt = wb_fmt[sname]
        max_row = min((ws.max_row or 0), header_rows + n_rows)
        max_col = ws.max_column or 0
        for c in range(1, max_col + 1):
            letter = get_column_letter(c)
            found = 0
            first = ""
            for r in range(header_rows + 1, max_row + 1):
                u = _cell_url(ws, r, c, ws_fmt)
                if u and _looks_like_link(u):
                    found += 1
                    if not first:
                        first = u[:50]
            if found:
                hits.append({"sheet": sname, "letter": letter, "count": found, "sample": first})
    wb_val.close()
    wb_fmt.close()
    return hits


def column_sample(path: str | Path, link_col, header_rows: int = 1,
                  sheet: Optional[str] = None, n: int = 15) -> list[str]:
    """Diagnostic: what values do we actually see in the chosen column?
    Scans up to `n` data rows and reports the first non-empty one."""
    wb_val = openpyxl.load_workbook(path, read_only=False, data_only=True)
    wb_fmt = openpyxl.load_workbook(path, read_only=False, data_only=False)
    ws = _get_sheet(wb_val, sheet)
    ws_fmt = _get_sheet(wb_fmt, sheet)
    col = _col_to_idx(link_col)
    vals = []
    first_nonempty_row = None
    for r in range(header_rows + 1, (ws.max_row or 0) + 1):
        cell = ws.cell(row=r, column=col)
        v = cell.value
        hl = cell.hyperlink
        t = getattr(hl, "target", None) if hl else None
        f = ws_fmt.cell(row=r, column=col).value
        furl = ""
        if isinstance(f, str):
            m = _URL_RE.search(f)
            if m:
                furl = m.group(0)
        disp = (f"val={v!r}" if v not in (None, "") else "val=(空)")
        if t:
            disp += f" | hyperlink={t!r}"
        if furl:
            disp += f" | formula_url={furl!r}"
        if first_nonempty_row is None and (v not in (None, "") or t or furl):
            first_nonempty_row = r
        vals.append(f"row{r}: {disp}")
        if len(vals) >= n:
            break
    if first_nonempty_row:
        vals.append(f"（该列第一个非空数据在第 {first_nonempty_row} 行）")
    wb_val.close()
    wb_fmt.close()
    return vals


def _is_rate_limit(msg: str) -> bool:
    """Classify an error message as an anti-crawler / rate-limit signal.

    These are the failures that mean 'slow down', so the batch loop should back
    off (lengthen the inter-note interval) instead of charging ahead.
    """
    m = (msg or "").lower()
    return any(
        k in m
        for k in (
            "风控", "fe-verify", "300012", "300013", "429",
            "频率", "限流", "rate", "too many", "abuse",
            "访问过于频繁", "操作过于频繁", "请求过于频繁",
        )
    )


def run_excel(
    excel_path: str | Path,
    engine: CaptureEngine,
    link_col,
    out_col,
    *,
    header_rows: int = 1,
    sheet: Optional[str] = None,
    mode: str = "note",
    selector: Optional[str] = None,
    keyword: Optional[str] = None,
    region: Optional[tuple] = None,
    out_dir: Optional[str | Path] = None,
    thumb_width: int = 320,
    show_stats: bool = True,
    log=print,
    on_progress=None,
    # ---- anti-ban / throughput knobs (tune for big batches) ----
    min_interval: float = 2.0,
    max_interval: float = 30.0,
    batch_size: int = 25,
    batch_pause: float = 30.0,
    max_consecutive_blocks: int = 8,
) -> Path:
    """Run the whole batch. Returns the path of the saved .xlsx with images embedded."""
    excel_path = Path(excel_path)
    out_dir = Path(out_dir) if out_dir else excel_path.parent / "xhs_shots_output"
    out_dir.mkdir(parents=True, exist_ok=True)

    rows = read_links(excel_path, link_col, header_rows, sheet)
    if not rows:
        diag = column_sample(excel_path, link_col, header_rows, sheet)
        sheetname = sheet or "(active)"
        try:
            # scan EVERY sheet, deep, for any link-like content
            link_cols = find_link_columns(excel_path, None, header_rows, all_sheets=True)
        except Exception:  # noqa: BLE001
            link_cols = []
        hint = ""
        if link_cols:
            lines = "\n      ".join(
                f"{c['sheet']}!{c['letter']}  ({c['count']}条, 示例: {c['sample']})"
                for c in link_cols[:8]
            )
            hint = "\n  · 自动扫描全表发现：以下位置其实含有链接，请改选 →\n      " + lines
        else:
            hint = "\n  · 已跨所有工作表、深扫约 300 行，仍未发现任何链接形态的单元格。" \
                   "可能：①链接在更靠下的行；②链接是图片/二维码而非文本网址；" \
                   "③文件由程序生成且公式结果未被 Excel 计算过（请先用 Excel 打开、Ctrl+` 重算后另存）。"
        raise CaptureError(
            "所选列没有读到任何链接。\n"
            f"  · 表：{sheetname}  列：{link_col}  跳过表头：{header_rows} 行\n"
            f"  · 该列前几行实际内容：\n      " + ("\n      ".join(diag) if diag else "（整列数据行为空）") +
            "\n  · 排查：①链接是否真的在这一列？②是否在其他 sheet？③表头行数是否设多/设少？"
            "④若链接是公式生成(=HYPERLINK/引用)，本工具已支持从公式里提取网址。"
            + hint
        )

    wb = openpyxl.load_workbook(excel_path)
    ws = _get_sheet(wb, sheet)
    out_c = _col_to_idx(out_col)
    # ensure header for output column (skip if it falls inside a merged cell,
    # e.g. the sheet's title row — writing a MergedCell raises)
    hdr_row = max(header_rows, 1)
    _hr, _hc = _resolve_anchor(ws, hdr_row, out_c)
    if ws.cell(row=_hr, column=_hc).value in (None, ""):
        ws.cell(row=_hr, column=_hc).value = "笔记截图"

    out_xlsx = excel_path.parent / f"{excel_path.stem}_with_shots.xlsx"

    # ---- anti-ban / throughput state ----
    # current_delay grows on rate-limit failures (exponential backoff) and
    # decays on successes. This keeps the batch from charging into a hard block
    # when XHS starts throttling the Cloud IP.
    current_delay = float(min_interval)
    consecutive_blocks = 0
    since_checkpoint = 0
    aborted = False

    total = len(rows)
    embedded_count = 0
    for i, (row, url) in enumerate(rows, start=1):
        # --- 链接合法性校验 ---
        # 单元格里可能是纯文本/数字/标题等非链接内容（read_links 的兜底逻辑会把任何
        # 非空值当链接返回）。这类直接跳过并提示原值，而不是让 page.goto 抛
        # "Cannot navigate to invalid URL" 把整行标成 error。
        if not _looks_like_link(url):
            _out_set(ws, row, out_c, f"⏭️ 非链接已跳过 (原值: {str(url)[:50]})")
            log(f"[{i}/{total}] row {row} -> skipped (非链接，原值: {str(url)[:50]})  {url}")
            time.sleep(0.1)
            continue
        # 严格校验 + 补全协议头：裸域名、带空格/换行的 URL、只有协议头没内容等
        # 非法情况都在此拦截并跳过，绝不让 page.goto 抛 "invalid URL"。
        nav_url = _normalize_and_validate_url(url)
        if not nav_url:
            _out_set(ws, row, out_c, f"⏭️ 链接格式无效已跳过 (原值: {str(url)[:50]})")
            log(f"[{i}/{total}] row {row} -> skipped (链接格式无效，原值: {str(url)[:50]})  {url}")
            time.sleep(0.1)
            continue
        name = sanitize_filename(nav_url)[:40]
        img_path = out_dir / f"row{row}_{name}.png"
        status = "ok"
        msg = ""
        try:
            engine.capture(
                nav_url, mode, selector=selector, keyword=keyword, region=region,
                out_path=img_path, show_stats=show_stats,
            )
            # --- verify image file ---
            if not img_path.exists():
                raise CaptureError("截图文件未生成（capture 未报错但文件不存在）")
            img_size = img_path.stat().st_size
            if img_size == 0:
                raise CaptureError(f"截图文件为空（0 字节）")
            if img_size < 500:
                raise CaptureError(f"截图文件异常小（{img_size} 字节），可能截到空白页")

            img = XLImage(str(img_path))
            scale = thumb_width / img.width if img.width else 1
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)

            anchor_cell = (lambda r, c: f"{get_column_letter(c)}{r}")(*_resolve_anchor(ws, row, out_c))
            ws.add_image(img, anchor_cell)

            # Set row height so the image is visible (at least the scaled height, in points)
            # 1 pt ≈ 1.33 px, use a reasonable minimum
            target_row_height = max(int(img.height / 1.33), 40)
            current_h = ws.row_dimensions[row].height or 15
            if target_row_height > current_h:
                ws.row_dimensions[row].height = target_row_height

            # Also set column width for the output column
            out_col_letter = get_column_letter(out_c)
            current_w = ws.column_dimensions[out_col_letter].width or 8
            if current_w < 20:
                ws.column_dimensions[out_col_letter].width = 25

            _out_set(ws, row, out_c, f"✅ 已插入 ({img.width}×{img.height}px)")
            embedded_count += 1
        except LoginWallError as e:
            status = "login_wall"
            msg = str(e)
            _out_set(ws, row, out_c, "⛔ 需登录")
        except CaptureError as e:
            msg = str(e)
            # A URL that Playwright still rejected as invalid should be a clean
            # skip, not a scary "browser error".
            if "invalid URL" in msg or "Invalid URL" in msg:
                status = "skipped"
                _out_set(ws, row, out_c, f"⏭️ 链接无法访问已跳过 (原值: {str(url)[:50]})")
                log(f"[{i}/{total}] row {row} -> skipped (链接无效)  {url}")
            else:
                status = "error"
                _out_set(ws, row, out_c, f"❌ {msg[:120]}")
        except Exception as e:  # noqa: BLE001
            msg = repr(e)
            if "invalid URL" in msg or "Invalid URL" in msg:
                status = "skipped"
                _out_set(ws, row, out_c, f"⏭️ 链接无法访问已跳过 (原值: {str(url)[:50]})")
                log(f"[{i}/{total}] row {row} -> skipped (链接无效)  {url}")
            else:
                status = "error"
                _out_set(ws, row, out_c, f"❌ {msg[:120]}")

        detail = f"  ({msg[:80]})" if msg else ""
        log(f"[{i}/{total}] row {row} -> {status}{detail}  {url}")
        if on_progress:
            on_progress(i, total, row, url, status, msg)

        # ---- adaptive throttle: decide the next inter-note delay ----
        if status == "ok":
            consecutive_blocks = 0
            current_delay = max(min_interval, current_delay * 0.7)  # decay back toward base
        elif status == "login_wall":
            consecutive_blocks = 0  # auth issue, not throttling — no backoff
        elif _is_rate_limit(msg):
            consecutive_blocks += 1
            current_delay = min(max_interval, current_delay * 2.0)  # exponential backoff
            log(f"  ⚠️ 疑似限流/风控，自动拉长间隔到 {current_delay:.1f}s（连续第 {consecutive_blocks} 次）")
            if consecutive_blocks >= max_consecutive_blocks:
                log(f"  🛑 连续 {consecutive_blocks} 次被拦截，疑似 Cloud IP 已被限流。"
                    f"停止批量以保护账号/IP，请降低频率、稍后重试。已处理 {i}/{total} 条。")
                aborted = True
                break
        else:
            # 其他错误（无效链接跳过、零字节等）：不是限流，不计入连续拦截
            consecutive_blocks = 0

        # ---- checkpoint + batch cooldown (resilience for big runs) ----
        since_checkpoint += 1
        if batch_size and since_checkpoint >= batch_size:
            try:
                wb.save(out_xlsx)
            except Exception:  # noqa: BLE001
                pass
            since_checkpoint = 0
            if i < total and not aborted:
                log(f"  💾 已 checkpoint（{i}/{total}），冷却 {batch_pause:.0f}s 让 IP 降温…")
                time.sleep(batch_pause)

        time.sleep(current_delay)  # polite + adaptive pause between requests

    wb.save(out_xlsx)
    tail = "（已中途停止：疑似被限流）" if aborted else ""
    log(f"\n批量完成{tail}：共 {total} 条，成功嵌入图片 {embedded_count} 张 → {out_xlsx}")
    return out_xlsx
